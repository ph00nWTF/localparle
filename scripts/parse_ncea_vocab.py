"""Parse NCEA French vocabulary Excel files into structured JSON.

Sources:
  FR to EN sheet  → main vocab list (flat alphabetical)
  Grammar and Structures sheet → grammar examples by category
  Expressions & Sample Sentences sheet (L2/L3) → example sentences

Each vocab entry has:
  french    — French word/phrase (article prepended for nouns)
  article   — le/la/l'/les/le/la (nouns only, else "")
  pos       — "noun" | "verb" | "other"
  english   — English gloss

Usage:
    uv run --with openpyxl python scripts/parse_ncea_vocab.py \
        --l1 <path/to/l1.xlsx> \
        --l2 <path/to/l2.xlsx> \
        --l3 <path/to/l3.xlsx> \
        --out data/curriculum/vocab_ncea.json
"""

import argparse
import json
import re
import sys


def get_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        sys.exit("openpyxl not found — run with: uv run --with openpyxl python ...")


ARTICLE_RE = re.compile(r"^(le/la|le|la|l'|les|un|une)\s*$", re.I)
ALPHA_RE = re.compile(r"^[A-Z]$")


def normalise_apos(s: str) -> str:
    """Normalise all apostrophe variants to straight ASCII apostrophe."""
    return s.replace("’", "'").replace("‘", "'").replace("ʼ", "'")


def cell(v) -> str:
    return normalise_apos(str(v or "").strip())


# ---------------------------------------------------------------------------
# FR to EN vocab sheet
# ---------------------------------------------------------------------------

def parse_fr_to_en(ws) -> list[dict]:
    """
    Columns: [alpha_marker, article, french, english, ...]
    Rows 1-4 are headers/description — skip them.
    """
    entries = []
    for row in ws.iter_rows(values_only=True):
        c0, c1, c2, c3 = (cell(row[i]) if i < len(row) else "" for i in range(4))

        # Skip header, description, and blank rows
        if not c2 or not c3:
            continue
        if c2 in ("French", "French ") or "Vocabulary List" in c3:
            continue
        if "Article" in c2 or "English" in c3:
            continue
        if len(c2) < 2:
            continue

        # Determine POS
        article = c1
        if ARTICLE_RE.match(article):
            pos = "noun"
            # Normalise apostrophe variants
            article = re.sub(r"[''']", "'", article).strip()
            french = f"{article} {c2}".strip()
        elif c3.startswith("to "):
            pos = "verb"
            article = ""
            french = c2
        else:
            pos = "other"
            article = ""
            french = c2

        entries.append({
            "french": french,
            "article": article,
            "pos": pos,
            "english": c3,
        })
    return entries


# ---------------------------------------------------------------------------
# Grammar and Structures sheet
# ---------------------------------------------------------------------------

def parse_grammar(ws) -> list[dict]:
    """
    Format:
      Category header row (single non-empty cell)
      Column header row ("French" | "English") [optional]
      Data rows: [concept/example_fr, example_en] or [concept, example_fr, example_en]
    """
    grammar = []
    current_category = None
    skip_next = False

    for row in ws.iter_rows(values_only=True):
        cells = [cell(c) for c in row if cell(c)]
        if not cells:
            continue

        # Single-cell row that isn't a header → category name
        if len(cells) == 1:
            val = cells[0]
            if val.startswith("Grammar") or val.startswith("'Grammar"):
                continue
            if "Vocabulary List" in val or "Words outside" in val:
                continue
            current_category = val
            skip_next = False
            continue

        if cells[0] in ("French", "English"):
            skip_next = True
            continue
        if skip_next:
            skip_next = False

        if current_category is None:
            continue

        # Data row
        if len(cells) == 2:
            grammar.append({
                "category": current_category,
                "concept": cells[0],
                "example_en": cells[1],
            })
        elif len(cells) >= 3:
            grammar.append({
                "category": current_category,
                "concept": cells[0],
                "example_fr": cells[1],
                "example_en": cells[2],
            })

    return grammar


# ---------------------------------------------------------------------------
# Expressions & Sample Sentences sheet
# ---------------------------------------------------------------------------

def parse_expressions(ws) -> list[dict]:
    """
    Format:
      Row 1: title
      Row 2: column headers [French, English, Example FR, Example EN]
      Data rows: [french, english, example_fr, example_en]
      Some rows are continuation rows (blank first two cells, example only)
    """
    expressions = []
    last_entry = None
    header_done = False

    for row in ws.iter_rows(values_only=True):
        # Data starts at col 1 (col 0 is always empty)
        cells = [cell(row[i]) if i < len(row) else "" for i in range(1, 5)]
        c0, c1, c2, c3 = cells

        if not any([c0, c1, c2, c3]):
            continue
        if "Expressions and Sample" in c0:
            continue
        if c0 in ("French", "English") or c1 in ("English", "Examples in French"):
            header_done = True
            continue
        if not header_done:
            continue

        if c0 and c1:
            # New entry
            entry = {"french": c0, "english": c1}
            if c2:
                entry["example_fr"] = c2
            if c3:
                entry["example_en"] = c3
            expressions.append(entry)
            last_entry = entry
        elif (c2 or c3) and last_entry is not None:
            # Continuation: additional examples for same entry
            if "examples" not in last_entry:
                # Promote first example to list
                first = {}
                if "example_fr" in last_entry:
                    first["fr"] = last_entry.pop("example_fr")
                if "example_en" in last_entry:
                    first["en"] = last_entry.pop("example_en")
                last_entry["examples"] = [first] if first else []
            extra = {}
            if c2:
                extra["fr"] = c2
            if c3:
                extra["en"] = c3
            if extra:
                last_entry["examples"].append(extra)

    return expressions


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

CEFR_APPROX = {"L1": "A1-A2", "L2": "B1", "L3": "B2"}


def parse_level(wb, level: str) -> dict:
    openpyxl = get_openpyxl()
    result: dict = {"cefr_approx": CEFR_APPROX[level]}

    result["vocab"] = parse_fr_to_en(wb["FR to EN"])

    if "Grammar and Structures" in wb.sheetnames:
        result["grammar"] = parse_grammar(wb["Grammar and Structures"])

    if "Expressions & Sample Sentences" in wb.sheetnames:
        result["expressions"] = parse_expressions(wb["Expressions & Sample Sentences"])

    counts = {
        "vocab": len(result["vocab"]),
        "nouns": sum(1 for e in result["vocab"] if e["pos"] == "noun"),
        "verbs": sum(1 for e in result["vocab"] if e["pos"] == "verb"),
        "other": sum(1 for e in result["vocab"] if e["pos"] == "other"),
        "grammar": len(result.get("grammar", [])),
        "expressions": len(result.get("expressions", [])),
    }
    return result, counts


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--l1", required=True)
    parser.add_argument("--l2", required=True)
    parser.add_argument("--l3", required=True)
    parser.add_argument("--out", default="data/curriculum/vocab_ncea.json")
    args = parser.parse_args()

    openpyxl = get_openpyxl()

    output = {
        "source": "NZALT — NCEA French Revised Vocabulary List 2026/2025 (Ministry of Education NZ)",
        "note": (
            "Level 1 ≈ CEFR A1-A2 (2026 list), Level 2 ≈ B1 (2026 list), Level 3 ≈ B2 (2025 list). "
            "Parsed from official Excel files. pos: noun|verb|other."
        ),
        "levels": {},
    }

    for level, path in [("L1", args.l1), ("L2", args.l2), ("L3", args.l3)]:
        print(f"Parsing {level} from {path}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        level_data, counts = parse_level(wb, level)
        for k, v in counts.items():
            print(f"  {k}: {v}")
        output["levels"][level] = level_data

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nWrote {args.out}")


if __name__ == "__main__":
    main()
